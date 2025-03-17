import os
import shutil
import time
from ipproxyclient import IpProxyClient
import openpyxl
from datetime import datetime
from jzscclient import JzscClient
from setting import Setting
from stateutil_yejijishuzhibiao import YjjszbStateUtil
from logutil import LogUtil


# 目录
g_config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'configs')
g_yjjszb_data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', '业绩技术指标')
os.makedirs(g_yjjszb_data_path, exist_ok=True)


# 移除无效字符
def remove_invalid_char(content):
    if content is None:
        return ''

    # 去除不可打印字符
    cleaned_str = ''.join(char for char in content if char.isprintable())
    return cleaned_str


def save_datas(datas):
    if len(datas) == 0:
        return

    # 根据最后一个数据获取表格文件名
    data_id = int(datas[-1].id)
    index = (data_id - 1) // Setting.get().yjjszb_count_per_file + 1
    excel_file_path = os.path.join(g_yjjszb_data_path, '业绩技术指标{}.xlsx'.format(index))
    if not os.path.exists(excel_file_path):
        src_excel_file = os.path.join(g_config_path, '业绩技术指标模板.xlsx')
        shutil.copyfile(src_excel_file, excel_file_path)

    while True:
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.worksheets[0]
            current_row = sheet.max_row + 1
            for row in range(len(datas)):
                data = datas[row]
                sheet.cell(current_row, 1, remove_invalid_char(data.id))
                sheet.cell(current_row, 2, remove_invalid_char(data.project_id))
                sheet.cell(current_row, 3, remove_invalid_char(data.zizibiaozhun))
                sheet.cell(current_row, 4, remove_invalid_char(data.begin_date))
                sheet.cell(current_row, 5, remove_invalid_char(data.end_date))
                sheet.cell(current_row, 6, remove_invalid_char(data.guimo_dengji))
                sheet.cell(current_row, 7, remove_invalid_char(data.data_level))
                sheet.cell(current_row, 11, remove_invalid_char(data.enterprise_name))
                for person in data.people:
                    sheet.cell(current_row, 8, remove_invalid_char(person.name))
                    sheet.cell(current_row, 9, remove_invalid_char(person.shenfenzheng_id))
                    sheet.cell(current_row, 10, remove_invalid_char(person.role))
                    current_row += 1
            workbook.save(excel_file_path)
            break
        except Exception as e:
            print('保存数据遇到问题：{}'.format(e))
            print('如果表格文件被打开，请先关闭，过30秒程序将继续尝试保存')
            time.sleep(30)


def main():
    print('启动业绩技术指标信息采集工具')
    jzsc_client = JzscClient()
    datas = []
    not_has_data_count = 0  # 连续没有数据的个数
    failed_count = 0  # 连续请求失败的次数
    failed_proxy_count = 0  # 连续失败代理IP数
    proxy_expire_time = 0
    begin_collect_id = max(Setting.get().yjjszb_begin_id, YjjszbStateUtil.get().next_collect_id)
    for collect_id in range(begin_collect_id, Setting.get().yjjszb_end_id):
        while True:
            time.sleep(Setting.get().yjjszb_collect_interval)

            if failed_proxy_count >= 20:
                print('连续{}个代理IP不能使用，退出程序'.format(failed_proxy_count))
                return

            # 过期前10秒开始换代理
            if proxy_expire_time == 0 or datetime.now().timestamp() + 10 >= proxy_expire_time or failed_count >= 5:
                failed_count = 0
                failed_proxy_count += 1
                proxy_client = IpProxyClient()
                proxy_client.link = Setting.get().proxy_ip_link
                while True:
                    time.sleep(1)
                    print('获取代理IP地址')
                    request_success, proxy, _ = proxy_client.extract_ip()
                    if not request_success:
                        continue
                    jzsc_client.proxies['http'] = 'http://{}:{}'.format(proxy[0], proxy[1])
                    jzsc_client.proxies['https'] = jzsc_client.proxies['http']
                    proxy_expire_time = proxy[2]
                    break

            print('开始采集第{}个业绩技术指标'.format(collect_id))
            request_success, data, _ = jzsc_client.get_jishuzhibiao(collect_id)
            if not request_success:
                failed_count += 1
                if jzsc_client.forbidden:
                    print('服务器禁止访问')
                continue

            failed_count = 0
            failed_proxy_count = 0
            if data is None:
                not_has_data_count += 1
            else:
                not_has_data_count = 0
                datas.append(data)
            break

        if data is None:
            continue

        while True:
            time.sleep(Setting.get().yjjszb_collect_interval)

            if failed_proxy_count >= 20:
                print('连续{}个代理IP不能使用，退出程序'.format(failed_proxy_count))
                return

            # 过期前10秒开始换代理
            if proxy_expire_time == 0 or datetime.now().timestamp() + 10 >= proxy_expire_time or failed_count >= 5:
                failed_count = 0
                failed_proxy_count += 1
                proxy_client = IpProxyClient()
                proxy_client.link = Setting.get().proxy_ip_link
                while True:
                    time.sleep(1)
                    print('获取代理IP地址')
                    request_success, proxy, _ = proxy_client.extract_ip()
                    if not request_success:
                        continue
                    jzsc_client.proxies['http'] = 'http://{}:{}'.format(proxy[0], proxy[1])
                    jzsc_client.proxies['https'] = jzsc_client.proxies['http']
                    proxy_expire_time = proxy[2]
                    break

            print('开始采集第{}个业绩技术指标的相关人员'.format(collect_id))
            request_success, people, _ = jzsc_client.get_xiangguanrenyuan(datas[-1].yejijilubianhao)
            if not request_success:
                failed_count += 1
                if jzsc_client.forbidden:
                    print('服务器禁止访问')
                continue

            failed_count = 0
            failed_proxy_count = 0
            datas[-1].people = people
            break

        # 采集结束或每采集到200条数据或需要换新文件的时候就保存
        is_finish = not_has_data_count >= 100
        if is_finish or len(datas) >= 200 or collect_id % Setting.get().yjjszb_count_per_file == 0:
            print('保存数据')
            save_datas(datas)
            print('保存数据完成')

            if len(datas) > 0:
                YjjszbStateUtil.get().update_next_collect_id(int(datas[-1].id)+1)
            datas = []

        if is_finish:
            print('采集完成')
            break


if __name__ == '__main__':
    LogUtil.file_name_prefix = 'yjjszb'
    LogUtil.enable()
    main()
