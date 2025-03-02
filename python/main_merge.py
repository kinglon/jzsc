import os
import shutil
import sys
import time

import openpyxl

from logutil import LogUtil


def do_merge(data_path, excel_template):
    excel_files = []
    for excel_file in os.listdir(data_path):
        if excel_file.find('合并') >= 0:
            continue
        excel_files.append(os.path.join(data_path, excel_file))

    if len(excel_files) == 0:
        print('没有表格需要合并')
        return

    excel_files.sort(key=os.path.getctime)
    merge_excel_file = os.path.join(data_path, '合并.xlsx')
    if os.path.exists(merge_excel_file):
        os.remove(merge_excel_file)
    shutil.copyfile(excel_template, merge_excel_file)

    merge_workbook = openpyxl.load_workbook(merge_excel_file)
    merge_sheet = merge_workbook.worksheets[0]
    current_row = 2
    for excel_file in excel_files:
        print('合并表格：{}'.format(os.path.split(excel_file)[1]))
        workbook = openpyxl.load_workbook(os.path.join(data_path, excel_file))
        sheet = workbook.worksheets[0]
        for row in range(2, sheet.max_row+1):
            column = 1
            for cell in sheet[row]:
                merge_sheet.cell(current_row, column, cell.value)
                column += 1
            current_row += 1
        workbook.close()

    while True:
        try:
            merge_workbook.save(merge_excel_file)
            break
        except Exception as e:
            print('保存数据遇到问题：{}'.format(e))
            print('如果表格文件被打开，请先关闭，过30秒程序将继续尝试保存')
            time.sleep(30)


def main():
    if len(sys.argv) < 2:
        raise RuntimeError('输入参数有误')

    # 1 合并竣工验收表格， 2 合并业绩技术指标表格
    data_type = sys.argv[1]
    if data_type != '1' and data_type != '2':
        raise RuntimeError('输入参数有误')

    data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'configs')
    if data_type == '1':
        jgys_data_path = os.path.join(data_path, '竣工验收')
        jgys_excel_template = os.path.join(config_path, '竣工验收模板.xlsx')
        do_merge(jgys_data_path, jgys_excel_template)
    elif data_type == '2':
        yjjszb_data_path = os.path.join(data_path, '业绩技术指标')
        yjjszb_excel_template = os.path.join(config_path, '业绩技术指标模板.xlsx')
        do_merge(yjjszb_data_path, yjjszb_excel_template)

    print('完成')


if __name__ == '__main__':
    LogUtil.file_name_prefix = 'merge'
    LogUtil.enable()
    main()
