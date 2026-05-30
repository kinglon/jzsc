import json
import os
import shutil
import time
from ipproxyclient import IpProxyClient
import openpyxl
from datetime import datetime
from jzscclient import JzscClient
from setting import Setting
from stateutil_yejijishuzhibiao import YjjszbStateUtil
from logutil import LogUtil


# 目录
g_config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'configs')
g_fake_data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', '虚假业绩')
os.makedirs(g_fake_data_path, exist_ok=True)
g_fake_state_file = os.path.join(g_fake_data_path, '虚假业绩采集状态.txt')
g_fake_state = {}

# 移除无效字符
def remove_invalid_char(content):
    if content is None:
        return ''

    # 去除不可打印字符
    cleaned_str = ''.join(char for char in content if char.isprintable())
    return cleaned_str

# 保存状态
def save_state():
    with open(g_fake_state_file, "w", encoding='utf-8') as file:
        json.dump(g_fake_state, file, indent=4, ensure_ascii=False)

def save_datas(datas):
    if len(datas) == 0:
        return

    excel_file_path = g_fake_state.get('saved_file_path', '')
    if len(excel_file_path) == 0:
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file_path = os.path.join(g_fake_data_path, f'虚假业绩_{current_time}.xlsx')
        g_fake_state['saved_file_path'] = excel_file_path
        save_state()

    if not os.path.exists(excel_file_path):
        src_excel_file = os.path.join(g_config_path, '虚假业绩模板.xlsx')
        shutil.copyfile(src_excel_file, excel_file_path)

    while True:
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.worksheets[0]
            current_row = sheet.max_row + 1
            for row in range(len(datas)):
                data = datas[row]
                sheet.cell(current_row, 1, remove_invalid_char(data['FAKE_NUM']))
                sheet.cell(current_row, 2, remove_invalid_char(data['PRJNUM']))
                sheet.cell(current_row, 3, remove_invalid_char(data['CORPNAME']))
                sheet.cell(current_row, 4, remove_invalid_char(data['PRJNAME']))
                sheet.cell(current_row, 5, remove_invalid_char(data['FAKETYPE']))
                mark_date = datetime.fromtimestamp(data['MARKDATE']/1000).strftime('%Y-%m-%d')
                sheet.cell(current_row, 6, mark_date)
                current_row += 1
            workbook.save(excel_file_path)
            break
        except Exception as e:
            print('保存数据遇到问题：{}'.format(e))
            print('如果表格文件被打开，请先关闭，过30秒程序将继续尝试保存')
            time.sleep(30)

def main():
    print('启动虚假业绩采集工具')

    # 加载采集状态
    if os.path.exists(g_fake_state_file):
        with open(g_fake_state_file, 'r', encoding='utf-8') as file:
            global g_fake_state
            g_fake_state = json.loads(file.read())

    jzsc_client = JzscClient()
    datas = []
    failed_count = 0  # 连续请求失败的次数
    failed_proxy_count = 0  # 连续失败代理IP数
    proxy_expire_time = 0

    # 逐页采集
    begin_page = g_fake_state.get('page', 0)
    for page in range(begin_page, 10000000):
        # 采集当前页数据直到成功，或者多次尝试失败直接退出程序
        while True:
            time.sleep(1)

            if failed_proxy_count >= 20:
                print('连续{}个代理IP不能使用，退出程序'.format(failed_proxy_count))
                return

            # 过期前10秒开始换代理
            if proxy_expire_time == 0 or datetime.now().timestamp() + 10 >= proxy_expire_time or failed_count >= 5:
                failed_count = 0
                failed_proxy_count += 1
                proxy_client = IpProxyClient()
                proxy_client.link = Setting.get().proxy_ip_link
                while True:
                    time.sleep(1)
                    print('获取代理IP地址')
                    request_success, proxy, _ = proxy_client.extract_ip()
                    if not request_success:
                        continue
                    jzsc_client.proxies['http'] = 'http://{}:{}'.format(proxy[0], proxy[1])
                    jzsc_client.proxies['https'] = jzsc_client.proxies['http']
                    proxy_expire_time = proxy[2]
                    break

            print(f'开始采集第{page+1}页虚假业绩')
            request_success, data, _ = jzsc_client.get_fake_yeji(page)
            if not request_success:
                failed_count += 1
                if jzsc_client.forbidden:
                    print('服务器禁止访问')
                continue

            failed_count = 0
            failed_proxy_count = 0
            break

        if len(data) == 0:
            print('已经采集到最后一页')
            break

        datas.extend(data)

        # 如果已经采集了100条，就先保存下来
        if len(datas) >= 100:
            print(f'保存数据：{len(datas)}条')
            save_datas(datas)
            datas = []
            print('保存数据完成')

            g_fake_state['page'] = page+1
            save_state()

    # 最后还有数据需要保存
    if len(datas) > 0:
        print(f'保存数据：{len(datas)}条')
        save_datas(datas)
        print('保存数据完成')

    # 删除状态文件
    if os.path.exists(g_fake_state_file):
        os.remove(g_fake_state_file)

    print('采集完成')

if __name__ == '__main__':
    LogUtil.file_name_prefix = 'fake'
    LogUtil.enable()
    main()
