import os
import shutil
import sys
import time
from datetime import datetime
import openpyxl

from ipproxyclient import IpProxyClient
from jzscclient import JzscClient
from setting import Setting
from stateutil_jungongyanshou import JgysStateUtil
from logutil import LogUtil

# 类型：1 竣工验收备案， 2 竣工验收
g_type = 1

# 目录
g_config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'configs')
g_jgys_data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', '竣工验收')
os.makedirs(g_jgys_data_path, exist_ok=True)
g_jgys_beian_data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', '竣工验收备案')
os.makedirs(g_jgys_beian_data_path, exist_ok=True)


# 移除无效字符
def remove_invalid_char(content):
    if content is None:
        return ''

    # 去除不可打印字符
    cleaned_str = ''.join(char for char in content if char.isprintable())
    return cleaned_str


def save_datas(datas):
    if len(datas) == 0:
        return

    # 根据最后一个数据获取表格文件名
    data_id = int(datas[-1].id)
    index = data_id // Setting.get().jgys_count_per_file + 1
    if g_type == 1:
        excel_file_path = os.path.join(g_jgys_beian_data_path, '竣工验收备案{}.xlsx'.format(index))
    else:
        excel_file_path = os.path.join(g_jgys_data_path, '竣工验收{}.xlsx'.format(index))
    if not os.path.exists(excel_file_path):
        src_excel_file = os.path.join(g_config_path, '竣工验收模板.xlsx')
        shutil.copyfile(src_excel_file, excel_file_path)

    while True:
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.worksheets[0]
            begin_row_index = sheet.max_row + 1
            for row in range(len(datas)):
                data = datas[row]
                current_row = begin_row_index + row
                sheet.cell(current_row, 1, remove_invalid_char(data.id))
                sheet.cell(current_row, 2, remove_invalid_char(data.project_id))
                sheet.cell(current_row, 3, remove_invalid_char(data.data_level))
                sheet.cell(current_row, 4, remove_invalid_char(data.shigong_license_id))
                sheet.cell(current_row, 5, remove_invalid_char(data.actual_use_money))
                sheet.cell(current_row, 6, remove_invalid_char(data.begin_date))
                sheet.cell(current_row, 7, remove_invalid_char(data.end_date))
                sheet.cell(current_row, 8, remove_invalid_char(data.data_source))
                sheet.cell(current_row, 9, remove_invalid_char(data.construct_scale))
                sheet.cell(current_row, 10, remove_invalid_char(data.remark))
            workbook.save(excel_file_path)
            break
        except Exception as e:
            print('保存数据遇到问题：{}'.format(e))
            print('如果表格文件被打开，请先关闭，过30秒程序将继续尝试保存')
            time.sleep(30)


def main():
    if g_type == 1:
        print('启动竣工验收备案信息采集工具')
    else:
        print('启动竣工验收信息采集工具')
    jzsc_client = JzscClient()
    datas = []
    not_has_data_count = 0  # 连续没有数据的个数
    failed_count = 0  # 连续请求失败的次数
    failed_proxy_count = 0  # 连续失败代理IP数
    proxy_expire_time = 0

    # 逆序采集
    if g_type == 1:
        begin_collect_id = Setting.get().jgys_beian_begin_id
        end_collect_id = min(Setting.get().jgys_beian_end_id, JgysStateUtil.get().next_beian_collect_id)
    else:
        begin_collect_id = Setting.get().jgys_begin_id
        end_collect_id = min(Setting.get().jgys_end_id, JgysStateUtil.get().next_collect_id)
    for collect_id in range(end_collect_id, begin_collect_id, -1):
        while True:
            time.sleep(Setting.get().jgys_collect_interval)

            if failed_proxy_count >= 20:
                print('连续{}个代理IP不能使用，退出程序'.format(failed_proxy_count))
                return

            # 过期前10秒开始换代理
            if proxy_expire_time == 0 or datetime.now().timestamp() + 10 >= proxy_expire_time or failed_count >= 5:
                failed_count = 0
                failed_proxy_count += 1
                proxy_client = IpProxyClient()
                proxy_client.link = Setting.get().proxy_ip_link
                while True:
                    time.sleep(1)
                    print('获取代理IP地址')
                    request_success, proxy, _ = proxy_client.extract_ip()
                    if not request_success:
                        continue
                    jzsc_client.proxies['http'] = 'http://{}:{}'.format(proxy[0], proxy[1])
                    jzsc_client.proxies['https'] = jzsc_client.proxies['http']
                    proxy_expire_time = proxy[2]
                    break

            print('开始采集第{}个'.format(collect_id))
            if g_type == 1:
                request_success, data, _ = jzsc_client.get_jungongyanshou_beian(collect_id)
            else:
                request_success, data, _ = jzsc_client.get_jungongyanshou(collect_id)
            if not request_success:
                failed_count += 1
                if jzsc_client.forbidden:
                    print('服务器禁止访问')
                continue

            failed_count = 0
            failed_proxy_count = 0
            if data is None:
                print('第{}个没有数据'.format(collect_id))
                not_has_data_count += 1
            else:
                not_has_data_count = 0
                datas.append(data)
            break

        # 采集结束或每采集到200条数据或需要换新文件的时候就保存
        is_finish = not_has_data_count >= 1000
        if is_finish or len(datas) >= 200 or collect_id % Setting.get().jgys_count_per_file == 0:
            print('保存数据')
            save_datas(datas)
            print('保存数据完成')

            if len(datas) > 0:
                if g_type == 1:
                    JgysStateUtil.get().update_next_beian_collect_id(int(datas[-1].id)-1)
                else:
                    JgysStateUtil.get().update_next_collect_id(int(datas[-1].id) - 1)
            datas = []

        if is_finish:
            print('采集完成')
            break
            
    # 最后还有数据没保存
    if len(datas) > 0:
        print('保存数据')
        save_datas(datas)
        print('保存数据完成')
        if g_type == 1:
            JgysStateUtil.get().update_next_beian_collect_id(int(datas[-1].id) - 1)
        else:
            JgysStateUtil.get().update_next_collect_id(int(datas[-1].id) - 1)
        print('采集完成')        


if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] == '2':
        g_type = 2
    else:
        g_type = 1

    LogUtil.file_name_prefix = 'jgys' if g_type == 2 else 'jgys_beian'
    LogUtil.enable()
    main()